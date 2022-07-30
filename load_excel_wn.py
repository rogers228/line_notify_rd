import time, datetime 
import openpyxl
import pandas as pd
import json
import config
import tool_db_yst

class Load_xls(): #讀取excel
    def __init__(self):
        self.file = config.excel_wn
        self.wb = openpyxl.load_workbook(self.file)
        self.sh = self.wb['待追蹤、修改']
        self.x_title = 1 #標題列位置
        self.x_data_statr = 2 # 資料開始
        self.x_data_over = 500 #資料結束
        self.load_df() # load data return self.df
        self.db = tool_db_yst.YST() 

    def load_df(self): # 產品資料
        sh = self.sh
        x_title = self.x_title #標題列位置
        x_data_statr = self.x_data_statr # 資料開始
        x_data_over  = self.x_data_over  #資料結束
        lis_title = []
        for i in range(1, sh.max_column+1):
            lis_title.append(sh.cell(x_title,i).value)

        # print(lis_title)
        lis_standard = {'品號': None,'品名': None,'變更後版次': None,'收件日': None,'最近檢查日期': None,
                        '單別': None,'單號': None,'提醒日期': None,'製程進度': None} # 合法欄位 excel第n欄

        for e in list(lis_standard.keys()):
            if e in lis_title:
                for i in range(1, sh.max_column+1):
                    if e == sh.cell(x_title, i).value:
                        lis_standard[e] = i
                        break
            else:
                self.err = {'is_err': True, 'message': f"缺少 '{e}' 欄位!"}
                print(self.err)
                return
        # print(lis_standard)

        lis_data = []
        for i in range(x_data_statr, x_data_over+1):
            row_data = []
            for key in list(lis_standard.keys()):
                myvalue = str(sh.cell(i, lis_standard[key]).value)
                if myvalue == 'None':
                    myvalue = ''
                row_data.append(myvalue)
                # row_data.append(str(sh.cell(i, lis_standard[key]).value))
            lis_data.append(row_data)
        # print(lis_data)
        df = pd.DataFrame(lis_data, columns = lis_standard) # 建立 DataFrame
        df['收件日'] =        pd.to_datetime(df['收件日'], format='%Y-%m-%d %H:%M:%S')
        df['最近檢查日期'] =  pd.to_datetime(df['最近檢查日期'], format='%Y-%m-%d %H:%M:%S')
        df['提醒日期'] =      pd.to_datetime(df['提醒日期'], format='%Y-%m-%d %H:%M:%S')
        df = df.fillna('') # 填充NaN為空白
        self.df = df

    def get_cgp(self): #已結案或已指定結案
        df = self.df
        df_w = df.loc[(df['品號'] != '') & (df['單別'] != '')] # 篩選

        lis_data = []
        dic_func = {'3': self.db.ger_ptd_kd, '5': self.db.ger_tad_kd}
        for i, r in df_w.iterrows():
            func = dic_func[str(r['單別'])[:1]]
            # print(func.__name__)
            cancel_no = func(r['單別'], r['單號'], r['品號'])  # 結案碼
            # print(r['單別'], r['單號'], r['品號'], f'結案碼:{cancel_no}')
            if cancel_no in ['y', 'Y']:
                lis_data.append([r['品號'], r['品名'], r['變更後版次']])

        df_y = pd.DataFrame(lis_data, columns = ['品號', '品名', '變更後版次']) # 建立 DataFrame
        return df_y

    def get_hhk(self): #已到提醒日期
        df = self.df
        df_w = df.loc[df['提醒日期'] != ''] # 篩選

        today = datetime.datetime.now()
        lis_data = []
        for i, r in df_w.iterrows():
            alarmday = r['提醒日期'].to_pydatetime()
            timeout_days = (today-alarmday).days # 超時天數
            # print(timeout_days)
            if timeout_days >= 0:
                lis_data.append([r['品號'], r['品名'], r['變更後版次'], r['提醒日期']])
            
        df_y = pd.DataFrame(lis_data, columns = ['品號', '品名', '變更後版次', '提醒日期']) # 建立 DataFrame
        return df_y


def test1():
    xls = Load_xls()
    # xls.test()
    df = xls.get_hhk()
    # df = xls.get_cgp()
    print(df)
    
    # xls.get_last_date()

if __name__ == '__main__':
    test1()
    print('ok')