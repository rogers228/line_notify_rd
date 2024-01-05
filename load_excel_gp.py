import time
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import json
import config

class Load_xls(): #讀取excel
    def __init__(self):
        self.file = config.sys_excel
        self.wb = openpyxl.load_workbook(self.file)
        self.sh = self.wb['車銑床回饋區']
        self.x_title = 1 #標題列位置
        self.x_data_statr = 2 # 資料開始
        self.x_data_over = 500 #資料結束
        self.load_df() # load data return self.df

    def load_df(self): # 產品資料
        sh = self.sh
        x_title = self.x_title #標題列位置
        x_data_statr = self.x_data_statr # 資料開始
        x_data_over  = self.x_data_over  #資料結束
        lis_title = []
        for i in range(1, sh.max_column+1):
            lis_title.append(sh.cell(x_title,i).value)

        # print(lis_title)
        lis_standard = {'日期': None,'提出人': None,'圖號(版別)': None,'問題': None,'完成日期': None,
                        '修改人員\n(改完簽名)': None} # 合法欄位 excel第n欄
        for e in list(lis_standard.keys()):
            if e in lis_title:
                for i in range(1, sh.max_column+1):
                    if e == sh.cell(x_title, i).value:
                        lis_standard[e] = i
                        break
            else:
                self.err = {'is_err': True, 'message': f"缺少 '{e}' 欄位!"}
                print(self.err)
                return
        # print(lis_standard)

        lis_data = []
        for i in range(x_data_statr, x_data_over+1):
            row_data = []
            for key in list(lis_standard.keys()):
                row_data.append(sh.cell(i, lis_standard[key]).value)
            lis_data.append(row_data)
        # print(lis_data)
        df = pd.DataFrame(lis_data, columns = lis_standard) # 建立 DataFrame
        df = df.fillna('') # 填充NaN為空白

        self.df = df

    def get_nup(self): # 尚未修改的資料
        df = self.df
        # print(df)
        df_w = df.loc[(df['日期'] != '') & (df['提出人'] != '') & (df['完成日期'] == '')] # 篩選
        df_g = df_w.copy() #複製
        df_g['日期'] =  pd.to_datetime(df_g['日期'], format='Y-%m-%d') # 時間格式化  轉時間
        df_g['日期'] = df_g['日期'].dt.strftime('%Y-%m-%d')  #格式化後 再轉回文字
        return df_g

    def getjson_nup(self):
        df = self.get_nup()
        # print(df) , date_format='iso'
        json_str = '{"data":' + df.to_json(orient="records") + '}'

        dic_data = json.loads(json_str)  # string-json variable to dict
        dic_data['lasttime'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

        json_s = json.dumps(dic_data, indent=4, ensure_ascii=False) # dict to format json
        return json_s

    def get_nup_count(self):
        df = self.get_nup()
        return len(df.index)

    def get_last_date(self): # 最後登入日期
        df = self.df
        df_w = df.loc[(df['日期'] != '') & (df['提出人'] != '') & (df['完成日期'] == '')] # 篩選
        # print(df_w)
        date = df_w["日期"].max().to_period('D')
        return date

    def get_first_date(self): # 最早登入日期
        df = self.df
        df_w = df.loc[(df['日期'] != '') & (df['提出人'] != '') & (df['完成日期'] == '')] # 篩選
        # print(df_w)
        date = df_w["日期"].min().to_period('D')
        return date

    def get_expired_dic(self): # 已經過期的最早登入日期的詳細資訊
        df = self.df
        df_w = df.loc[(df['日期'] != '') & (df['提出人'] != '') & (df['完成日期'] == '')] # 篩選
        # print(df_w)
        # print(df_w.columns)
        df_s = df_w.copy()
        df_s['days'] = (datetime.now() - df_w['日期']).dt.days # 距今日天數
        df_s = df_s.sort_values(by='days', ascending=False)    # 排序
        # print(df_w)
        dic = {}
        e = df_s.iloc[0]
        if e['days'] > 7: #過期7日
            dic['expired'] = True
            date = e['日期']
            mydate_datetime2Str = date.strftime('%Y-%m-%d')
            # dic['info'] = f"{e['日期'].strftime('%Y-%m-%d')} 由{e['提出人']}登記，{e['圖號(版別)']} 尚未修改已超過7日!"
            dic['info'] = f"由{e['提出人']}登記，{e['圖號(版別)']} 尚未修改已超過7日!"
        else:
            dic['expired'] = False
            dic['info'] = ''
        return dic

def test1():
    xls = Load_xls()
    # print(xls.get_nup_count())
    # print(xls.get_last_date())
    # print(xls.get_first_date())
    print(xls.get_expired_dic())


if __name__ == '__main__':
    test1()
    print('ok')