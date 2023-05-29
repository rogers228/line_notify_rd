import time, datetime 
import openpyxl
import pandas as pd
import json
import config
import tool_db_yst

class Load_xls(): #讀取excel
    def __init__(self):
        self.file = config.excel_wn
        self.wb = openpyxl.load_workbook(self.file)
        self.sh = self.wb['製令追蹤']
        self.x_title = 1 #標題列位置
        self.x_data_statr = 2 # 資料開始
        self.x_data_over = 500 #資料結束
        self.load_df() # load data return self.df
        self.db = tool_db_yst.YST() 

    def load_df(self): # 產品資料
        sh = self.sh
        x_title = self.x_title #標題列位置
        x_data_statr = self.x_data_statr # 資料開始
        x_data_over  = self.x_data_over  #資料結束
        lis_title = []
        for i in range(1, sh.max_column+1):
            lis_title.append(sh.cell(x_title,i).value)

        # print(lis_title)

        lis_standard = {'製令單別': None,'單號': None,'品號': None, '品名': None, '規格': None,
            '開單日期': None, '預計數量': None,'最近檢查日期': None, '提醒日期': None} # 合法欄位 excel第n欄
        # 提醒日期 使用者輸入  有超過提醒日期才會檢查
        # 最近檢查日期 使用者輸入  紀錄手動檢查的日期
        for e in list(lis_standard.keys()):
            if e in lis_title:
                for i in range(1, sh.max_column+1):
                    if e == sh.cell(x_title, i).value:
                        lis_standard[e] = i
                        break
            else:
                self.err = {'is_err': True, 'message': f"缺少 '{e}' 欄位!"}
                print(self.err)
                return
        # print(lis_standard)

        lis_data = []
        for i in range(x_data_statr, x_data_over+1):
            row_data = []
            for key in list(lis_standard.keys()):
                myvalue = str(sh.cell(i, lis_standard[key]).value)
                if myvalue == 'None':
                    myvalue = ''
                row_data.append(myvalue)
                # row_data.append(str(sh.cell(i, lis_standard[key]).value))
            lis_data.append(row_data)
        # print(lis_data)
        df = pd.DataFrame(lis_data, columns = lis_standard) # 建立 DataFrame
        df['最近檢查日期'] =  pd.to_datetime(df['最近檢查日期'], format='%Y-%m-%d %H:%M:%S')
        df['提醒日期'] =      pd.to_datetime(df['提醒日期'], format='%Y-%m-%d %H:%M:%S')
        df = df.fillna('') # 填充NaN為空白
        self.df = df

    def get_moy(self):
        # 超過提醒日期，進行檢查，不論是否結案
        df = self.df
        df_w = df.loc[(df['製令單別'] != '') & (df['單號'] != '') & (df['品號'] != '') & (df['提醒日期'] != '')] # 篩選
        # 提醒日期 使用者輸入  有超過提醒日期才會檢查    
        # print(df_w)
        today = datetime.datetime.now()
        lis_data = []
        for i, r in df_w.iterrows():
            alarmday = r['提醒日期'].to_pydatetime()
            timeout_days = (today-alarmday).days # 超時天數
            if timeout_days >= 0:
                cancel_no = self.db.ger_tad_kd(r['製令單別'], r['單號'], r['品號'])  # 製令結案碼 1.未生產,2.已發料,3.生產中,Y.已完工,y.指定完工
                lis_data.append([r['製令單別'],r['單號'], r['品號'], r['品名'], r['規格'], r['開單日期'], r['預計數量'], cancel_no])
        df_y = pd.DataFrame(lis_data, columns = ['製令單別','單號','品號', '品名', '規格','開單日期','預計數量','結案碼']) # 建立 DataFrame
        return df_y

def test1():
    xls = Load_xls()
    df = xls.get_moy()
    print(df)

if __name__ == '__main__':
    test1()
    print('ok')
